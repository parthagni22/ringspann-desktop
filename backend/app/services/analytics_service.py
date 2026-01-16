"""
Analytics Service
Business logic and database queries for analytics
"""
from warnings import filters
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, case, extract
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from app.utils.logger import setup_logger

logger = setup_logger()
import json
from app.models import Project, CommercialQuotation, TechnicalQuotation
from app.models.analytics_models import (
    AnalyticsFilters,
    ProductAnalyticsResponse,
    FinanceAnalyticsResponse,
    CustomerAnalyticsResponse,
    CombinedInsightsResponse,
    KPICard,
    ProductData,
    CustomerData,
    ChartDataPoint,
    TimeSeriesPoint
)


# Product type mapping
PART_TYPE_MAPPING = {
    '1': 'Brake Quotation',
    '2': 'Backstop Quotation',
    '3': 'Couple and Torque Limiter',
    '4': 'Locking Element for Conveyor',
    '5': 'Over Running Clutch',
    'Brake Quotation': 'Brake Quotation',
    'Backstop Quotation': 'Backstop Quotation',
    'Couple and Torque Limiter': 'Couple and Torque Limiter',
    'Locking Element for Conveyor': 'Locking Element for Conveyor',
    'Over Running Clutch': 'Over Running Clutch'
}


class AnalyticsService:
    """Service for analytics calculations and queries"""
    
    def __init__(self, db: Session):
        self.db = db
    
    # ========================================================================
    # UTILITY METHODS
    # ========================================================================
    
    def apply_date_filter(self, query, model, filters: AnalyticsFilters):
        """Apply date filtering to query"""
        if filters.date_filter == "today":
            start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            query = query.filter(model.created_at >= start_date)

        elif filters.date_filter == "custom":
            if filters.start_date:
                start = datetime.strptime(filters.start_date, "%Y-%m-%d")
                query = query.filter(model.created_at >= start)
            if filters.end_date:
                end = datetime.strptime(filters.end_date, "%Y-%m-%d")
                end = end.replace(hour=23, minute=59, second=59)
                query = query.filter(model.created_at <= end)

        return query
    
    def apply_status_filter(self, query, filters: AnalyticsFilters):
        """Apply quote status filter"""
        if filters.quote_status and filters.quote_status != "all":
            query = query.filter(Project.quote_status == filters.quote_status)
        return query
    
    def apply_customer_filter(self, query, filters: AnalyticsFilters):
        """Apply customer filter"""
        if filters.customer and filters.customer != "all":
            query = query.filter(Project.customer_name == filters.customer)
        return query
    
    def get_part_type_name(self, part_type: str) -> str:
        """Convert part type code to name"""
        return PART_TYPE_MAPPING.get(part_type, part_type)

    def get_part_type_code(self, product_name: str) -> str:
        """Convert product name to part type code"""
        # Reverse mapping from product name to code
        reverse_map = {v: k for k, v in PART_TYPE_MAPPING.items() if k.isdigit()}
        return reverse_map.get(product_name, product_name)

    def parse_commercial_items(self, items_json: str) -> List[Dict[str, Any]]:
        """Parse commercial quotation items JSON"""
        try:
            if not items_json:
                return []
            items = json.loads(items_json)
            return items if isinstance(items, list) else []
        except (json.JSONDecodeError, TypeError):
            return []
    
    def calculate_change_percent(self, current: float, previous: float) -> tuple:
        """Calculate percentage change and direction"""
        if previous == 0:
            return None, "neutral"
        
        change = ((current - previous) / previous) * 100
        direction = "up" if change > 0 else "down" if change < 0 else "neutral"
        return round(change, 2), direction
    
    # ========================================================================
    # PRODUCT ANALYTICS
    # ========================================================================
    
    def get_product_analytics(self, filters: AnalyticsFilters) -> ProductAnalyticsResponse:
        """Get complete product analytics"""
        
        # Get all required data
        quotes_by_product = self.get_quotes_by_product(filters)
        revenue_by_product = self.get_revenue_by_product(filters)
        product_trend = self.get_product_trend(filters)
        status_breakdown = self.get_product_status_breakdown(filters)
        
        # Calculate KPIs
        total_quotes = sum(p['quote_count'] for p in revenue_by_product)
        total_revenue = sum(p['revenue'] for p in revenue_by_product)
        
        # Find most quoted product
        most_quoted = max(quotes_by_product, key=lambda x: x['quote_count']) if quotes_by_product else None
        
        # Calculate average quote value per product
        avg_values = []
        for product in revenue_by_product:
            if product['quote_count'] > 0:
                avg_values.append(product['revenue'] / product['quote_count'])
        overall_avg = sum(avg_values) / len(avg_values) if avg_values else 0
        
        kpis = {
            "total_quotes": KPICard(
                label="Total Quotes",
                value=total_quotes,
                format_type="number"
            ),
            "total_revenue": KPICard(
                label="Total Revenue",
                value=round(total_revenue, 2),
                format_type="currency"
            ),
            "avg_quote_value": KPICard(
                label="Average Quote Value",
                value=round(overall_avg, 2),
                format_type="currency"
            ),
            "most_quoted_product": KPICard(
                label="Most Quoted Product",
                value=most_quoted['product_type'] if most_quoted else "N/A",
                format_type="text"
            ),
            "product_count": KPICard(
                label="Active Products",
                value=len(quotes_by_product),
                format_type="number"
            )
        }
        
        return ProductAnalyticsResponse(
            kpis=kpis,
            quote_distribution=revenue_by_product,
            revenue_contribution=revenue_by_product,
            product_trend=product_trend,
            status_breakdown=status_breakdown,
            filters_applied=filters.model_dump(),
            data_timestamp=datetime.now(),
            total_records=total_quotes
        )
    
    def get_quotes_by_product(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get quote count by product type"""
        query = self.db.query(
            TechnicalQuotation.part_type,
            func.count(TechnicalQuotation.id).label('quote_count')
        ).join(
            Project, TechnicalQuotation.quotation_number == Project.quotation_number
        )

        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        # Apply product_type filter
        if filters.product_type and filters.product_type != "all":
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        query = query.group_by(TechnicalQuotation.part_type)
    
        results = query.all()
        total = sum(r.quote_count for r in results)
    
        return [
        {
            "product_type": self.get_part_type_name(r.part_type),
            "quote_count": r.quote_count,
            "revenue": 0.0,  # Add this
            "avg_value": 0.0,  # Add this
            "percentage": round((r.quote_count / total * 100), 2) if total > 0 else 0
        }
        for r in results
    ]
    
    def get_revenue_by_product(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get revenue breakdown by product type"""

        # Get all commercial quotations with filters
        query = self.db.query(
            CommercialQuotation.quotation_number,
            CommercialQuotation.items,
            CommercialQuotation.subtotal,
            TechnicalQuotation.part_type
        ).join(
            Project, CommercialQuotation.quotation_number == Project.quotation_number
        ).outerjoin(
            TechnicalQuotation, CommercialQuotation.quotation_number == TechnicalQuotation.quotation_number
        )

        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        # Apply product_type filter
        if filters.product_type and filters.product_type != "all":
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        quotations = query.all()
        
        # Parse items and aggregate by product
        product_revenue = {}
        
        for quot in quotations:
            items = self.parse_commercial_items(quot.items)
            
            for item in items:
                product_name = item.get('description', 'Unknown')
                amount = float(item.get('amount', 0))
                
                if product_name not in product_revenue:
                    product_revenue[product_name] = {
                        'revenue': 0,
                        'quote_count': 0
                    }
                
                product_revenue[product_name]['revenue'] += amount
                product_revenue[product_name]['quote_count'] += 1
        
        # Convert to list and calculate percentages
        total_revenue = sum(p['revenue'] for p in product_revenue.values())
        
        result = []
        for product_name, data in product_revenue.items():
            avg_value = data['revenue'] / data['quote_count'] if data['quote_count'] > 0 else 0
            percentage = (data['revenue'] / total_revenue * 100) if total_revenue > 0 else 0
            
            result.append({
                "product_type": product_name,
                "revenue": round(data['revenue'], 2),
                "quote_count": data['quote_count'],
                "avg_value": round(avg_value, 2),
                "percentage": round(percentage, 2)
            })
        
        # Sort by revenue descending
        result.sort(key=lambda x: x['revenue'], reverse=True)
        
        return result
    
    def get_product_trend(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get product quotes trend over time"""

        # Determine grouping (monthly or weekly based on date range)
        query = self.db.query(
            TechnicalQuotation.part_type,
            func.strftime('%Y-%m', Project.created_at).label('period'),
            func.count(TechnicalQuotation.id).label('quote_count')
        ).join(
            Project, TechnicalQuotation.quotation_number == Project.quotation_number
        )

        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        if filters.product_type and filters.product_type != "all":
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)
        
        query = query.group_by(TechnicalQuotation.part_type, 'period').order_by('period')
        
        results = query.all()
        
        # Organize data by period with all products
        trend_data = {}
        products = set()
        
        for r in results:
            period = r.period
            product = self.get_part_type_name(r.part_type)
            products.add(product)
            
            if period not in trend_data:
                trend_data[period] = {'period': period}
            
            trend_data[period][product] = r.quote_count
        
        # Fill in missing products with 0
        for period_data in trend_data.values():
            for product in products:
                if product not in period_data:
                    period_data[product] = 0
        
        # Convert to list
        result = sorted(trend_data.values(), key=lambda x: x['period'])
        
        return result
    
    def get_product_status_breakdown(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get product quote status breakdown"""

        query = self.db.query(
            TechnicalQuotation.part_type,
            Project.quote_status,
            func.count(TechnicalQuotation.id).label('count')
        ).join(
            Project, TechnicalQuotation.quotation_number == Project.quotation_number
        )

        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        # Apply product_type filter
        if filters.product_type and filters.product_type != "all":
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        query = query.group_by(TechnicalQuotation.part_type, Project.quote_status)
        
        results = query.all()
        
        # Organize by product
        breakdown = {}
        for r in results:
            product = self.get_part_type_name(r.part_type)
            
            if product not in breakdown:
                breakdown[product] = {
                    'product_type': product,
                    'Budgetary': 0,
                    'Active': 0,
                    'Lost': 0,
                    'Won': 0,
                    'total': 0
                }
            
            breakdown[product][r.quote_status] = r.count
            breakdown[product]['total'] += r.count
        
        return list(breakdown.values())
    
    # Continued in next part...
    
    """
    Analytics Service - Part 2
        Finance, Customer, and Combined Insights
    """
    
    # ========================================================================
    # FINANCE ANALYTICS
    # ========================================================================
    
    def get_finance_analytics(self, filters: AnalyticsFilters):
        """Get complete finance analytics"""
        
        # Query WITHOUT items column to avoid JSON parsing errors
        query = self.db.query(
            CommercialQuotation.id,
            CommercialQuotation.quotation_number,
            CommercialQuotation.total_amount,
            CommercialQuotation.created_at,
            CommercialQuotation.updated_at,
            Project.quote_status,
            Project.customer_name,
            Project.created_at.label('project_created_at')
        ).join(
            Project, CommercialQuotation.quotation_number == Project.quotation_number
        )
        
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)
        
        results = query.all()
        
        # Calculate KPIs from results
        total_quoted_value = sum(float(r.total_amount or 0) for r in results)
        total_quotes = len(results)
        avg_quote_value = total_quoted_value / total_quotes if total_quotes else 0
        
        # Get status breakdown for value by status
        revenue_by_status_data = self.get_revenue_by_status(filters)
        
        # Get product revenue
        product_revenue = self.get_revenue_by_product(filters)
        top_product_revenue = max(product_revenue, key=lambda x: x['revenue']) if product_revenue else None
        
        kpis = {
            "total_quoted_value": {
                "label": "Total Quoted Value",
                "value": round(total_quoted_value, 2),
                "change_percent": None,
                "change_direction": "neutral",
                "format_type": "currency"
            },
            "total_quotes": {
                "label": "Total Quotes",
                "value": total_quotes,
                "change_percent": None,
                "change_direction": "neutral",
                "format_type": "number"
            },
            "avg_quote_value": {
                "label": "Average Quote Value",
                "value": round(avg_quote_value, 2),
                "change_percent": None,
                "change_direction": "neutral",
                "format_type": "currency"
            },
            "top_product": {
                "label": "Top Revenue Product",
                "value": top_product_revenue['product_type'] if top_product_revenue else "N/A",
                "change_percent": None,
                "change_direction": "neutral",
                "format_type": "text"
            },
            "top_product_revenue": {
                "label": "Top Product Revenue",
                "value": round(top_product_revenue['revenue'], 2) if top_product_revenue else 0,
                "change_percent": None,
                "change_direction": "neutral",
                "format_type": "currency"
            }
        }
        
        # Get charts data
        monthly_trend = self.get_monthly_revenue_trend(filters)
        value_distribution = self.get_quote_value_distribution(filters)
        inquiry_timeline = self.get_inquiry_timeline(filters)
        
        return {
            "kpis": kpis,
            "revenue_by_status": revenue_by_status_data,
            "monthly_trend": monthly_trend,
            "product_revenue": product_revenue,
            "value_distribution": value_distribution,
            "inquiry_timeline": inquiry_timeline,
            "filters_applied": filters.model_dump(),
            "data_timestamp": datetime.now().isoformat(),
            "total_records": total_quotes
        }
    
    def get_revenue_by_status(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get revenue breakdown by quote status"""
        
        query = self.db.query(
            Project.quote_status,
            func.count(CommercialQuotation.id).label('quote_count'),
            func.sum(CommercialQuotation.total_amount).label('total_revenue'),
            func.avg(CommercialQuotation.total_amount).label('avg_revenue')
        ).join(
            CommercialQuotation, Project.quotation_number == CommercialQuotation.quotation_number
        )
        
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_customer_filter(query, filters)

        if filters.product_type and filters.product_type != "all":
            query = query.join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            )
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        query = query.group_by(Project.quote_status)
        
        results = query.all()
        
        return [
            {
                "label": r.quote_status,
                "value": float(r.total_revenue or 0),
                "metadata": {
                    "quote_count": r.quote_count,
                    "avg_revenue": float(r.avg_revenue or 0)
                }
            }
            for r in results
        ]
    
    def get_monthly_revenue_trend(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get monthly revenue trend"""

        query = self.db.query(
            func.strftime('%Y-%m', Project.created_at).label('month'),
            func.count(CommercialQuotation.id).label('quote_count'),
            func.sum(CommercialQuotation.total_amount).label('total_revenue'),
            func.avg(CommercialQuotation.total_amount).label('avg_revenue')
        ).join(
            CommercialQuotation, Project.quotation_number == CommercialQuotation.quotation_number
        )

        # Apply ALL filters
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        # Apply product_type filter
        if filters.product_type and filters.product_type != "all":
            query = query.join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            )
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        query = query.group_by('month').order_by('month')

        results = query.all()
        
        return [
            {
                "date": r.month,
                "value": float(r.total_revenue or 0),
                "label": r.month,
                "metadata": {
                    "quote_count": r.quote_count,
                    "avg_revenue": float(r.avg_revenue or 0)
                }
            }
            for r in results
        ]
    
    def get_quote_value_distribution(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get distribution of quote values (histogram data)"""

        query = self.db.query(
            CommercialQuotation.total_amount
        ).join(
            Project, CommercialQuotation.quotation_number == Project.quotation_number
        )

        # Apply ALL filters
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        # Apply product_type filter
        if filters.product_type and filters.product_type != "all":
            query = query.join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            )
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        amounts = [r.total_amount for r in query.all() if r.total_amount]
        
        if not amounts:
            return []
        
        # Create bins
        min_val = min(amounts)
        max_val = max(amounts)
        bin_count = 10
        bin_size = (max_val - min_val) / bin_count if max_val > min_val else 1
        
        bins = {}
        for amount in amounts:
            bin_index = int((amount - min_val) / bin_size) if bin_size > 0 else 0
            bin_index = min(bin_index, bin_count - 1)
            
            bin_range = f"${int(min_val + bin_index * bin_size)}-${int(min_val + (bin_index + 1) * bin_size)}"
            
            if bin_range not in bins:
                bins[bin_range] = 0
            bins[bin_range] += 1
        
        return [
            {"range": k, "count": v}
            for k, v in sorted(bins.items())
        ]
    
    def get_inquiry_timeline(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get inquiry to quotation timeline"""

        # Use Project.created_at since inquiry_date doesn't exist
        query = self.db.query(
            Project.created_at,
            CommercialQuotation.quotation_number
        ).join(
            CommercialQuotation, Project.quotation_number == CommercialQuotation.quotation_number
        ).filter(
            Project.created_at.isnot(None)
        )

        # Apply ALL filters
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        # Apply product_type filter
        if filters.product_type and filters.product_type != "all":
            query = query.join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            )
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        results = query.all()
        
        # Group by month for timeline
        from collections import defaultdict
        monthly_counts = defaultdict(int)
        
        for r in results:
            if r.created_at:
                month_key = r.created_at.strftime("%Y-%m")
                monthly_counts[month_key] += 1
        
        timeline_data = [
            {
                "month": month,
                "count": count
            }
            for month, count in sorted(monthly_counts.items())
        ]
        
        return timeline_data
    
    # ========================================================================
    # CUSTOMER ANALYTICS
    # ========================================================================
    
    def get_customer_analytics(self, filters: AnalyticsFilters) -> CustomerAnalyticsResponse:
        """Get complete customer analytics"""

        # Get customer summary
        query = self.db.query(
            Project.customer_name,
            func.count(Project.id).label('quote_count'),
            func.max(Project.created_at).label('last_quote_date')
        )

        # Apply product_type filter if needed
        if filters.product_type and filters.product_type != "all":
            query = query.join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            )
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        query = query.group_by(Project.customer_name)

        # Apply other filters
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        customers = query.all()
        
        total_customers = len(customers)
        total_quotes = sum(c.quote_count for c in customers)
        
        # Get revenue data
        top_by_revenue = self.get_top_customers(filters, "revenue", 10)
        total_revenue = sum(c['revenue'] for c in top_by_revenue)
        avg_revenue_per_customer = total_revenue / total_customers if total_customers > 0 else 0
        
        # New vs Repeat customers
        new_customers = len([c for c in customers if c.quote_count == 1])
        repeat_customers = len([c for c in customers if c.quote_count > 1])
        
        kpis = {
            "total_customers": KPICard(
                label="Total Customers",
                value=total_customers,
                format_type="number"
            ),
            "total_revenue": KPICard(
                label="Total Revenue",
                value=round(total_revenue, 2),
                format_type="currency"
            ),
            "avg_revenue_per_customer": KPICard(
                label="Avg Revenue per Customer",
                value=round(avg_revenue_per_customer, 2),
                format_type="currency"
            ),
            "new_customers": KPICard(
                label="New Customers",
                value=new_customers,
                format_type="number"
            ),
            "repeat_customers": KPICard(
                label="Repeat Customers",
                value=repeat_customers,
                format_type="number"
            )
        }
        
        # Get charts data
        top_by_count = self.get_top_customers(filters, "quote_count", 10)
        status_breakdown = self.get_customer_status_breakdown(filters, 10)
        activity_timeline = self.get_customer_activity_timeline(filters)
        
        new_vs_repeat = {
            "new": new_customers,
            "repeat": repeat_customers,
            "total": total_customers
        }
        
        return CustomerAnalyticsResponse(
            kpis=kpis,
            top_customers_by_count=top_by_count,
            top_customers_by_revenue=top_by_revenue,
            customer_status_breakdown=status_breakdown,
            activity_timeline=activity_timeline,
            new_vs_repeat=new_vs_repeat,
            filters_applied=filters.model_dump(),
            data_timestamp=datetime.now(),
            total_records=total_customers
        )
    
    def get_top_customers(self, filters: AnalyticsFilters, sort_by: str, limit: int) -> List[Dict[str, Any]]:
        """Get top customers by revenue or quote count"""
        
        query = self.db.query(
            Project.customer_name,
            func.count(Project.id).label('quote_count'),
            func.coalesce(func.sum(CommercialQuotation.total_amount), 0).label('total_revenue'),
            func.max(Project.created_at).label('last_quote_date')
        ).outerjoin(
            CommercialQuotation, Project.quotation_number == CommercialQuotation.quotation_number
        )
        
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        if filters.product_type and filters.product_type != "all":
            query = query.join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            )
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        query = query.group_by(Project.customer_name)
        
        if sort_by == "revenue":
            query = query.order_by(func.sum(CommercialQuotation.total_amount).desc())
        else:
            query = query.order_by(func.count(Project.id).desc())
        
        query = query.limit(limit)
        
        results = query.all()
        
        result_list = []
        for r in results:
            revenue = float(r.total_revenue or 0)
            avg_deal = revenue / r.quote_count if r.quote_count > 0 else 0
            
            result_list.append({
                "customer_name": r.customer_name,
                "quote_count": r.quote_count,
                "revenue": round(revenue, 2),
                "avg_deal_size": round(avg_deal, 2),
                "last_quote_date": r.last_quote_date.strftime("%Y-%m-%d") if r.last_quote_date else None
            })
        
        return result_list
    
    def get_customer_status_breakdown(self, filters: AnalyticsFilters, limit: int) -> List[Dict[str, Any]]:
        """Get quote status breakdown per customer"""

        query = self.db.query(
            Project.customer_name,
            Project.quote_status,
            func.count(Project.id).label('count')
        )

        # Apply ALL filters
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        if filters.product_type and filters.product_type != "all":
            query = query.join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            )
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        query = query.group_by(Project.customer_name, Project.quote_status)
        
        results = query.all()
        
        # Organize by customer
        breakdown = {}
        for r in results:
            if r.customer_name not in breakdown:
                breakdown[r.customer_name] = {
                    'customer_name': r.customer_name,
                    'Budgetary': 0,
                    'Active': 0,
                    'Lost': 0,
                    'Won': 0,
                    'total': 0
                }
            
            breakdown[r.customer_name][r.quote_status] = r.count
            breakdown[r.customer_name]['total'] += r.count
        
        # Sort by total and limit
        sorted_customers = sorted(breakdown.values(), key=lambda x: x['total'], reverse=True)
        return sorted_customers[:limit]
    
    def get_customer_activity_timeline(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get customer activity timeline"""

        query = self.db.query(
            Project.customer_name,
            Project.quotation_number,
            Project.created_at,
            Project.quote_status
        )

        # Apply ALL filters
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        query = self.apply_customer_filter(query, filters)

        if filters.product_type and filters.product_type != "all":
            query = query.join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            )
            part_code = self.get_part_type_code(filters.product_type)
            query = query.filter(TechnicalQuotation.part_type == part_code)

        query = query.order_by(Project.created_at.desc())

        results = query.all()
        
        return [
            {
                "customer_name": r.customer_name,
                "quotation_number": r.quotation_number,
                "date": r.created_at.strftime("%Y-%m-%d"),
                "status": r.quote_status
            }
            for r in results
        ]
    
    # Continued in part 3...
    """
Analytics Service - Part 3
Combined Insights and Export functionality
"""
    
    # ========================================================================
    # COMBINED INSIGHTS
    # ========================================================================
    
    def get_combined_insights(self, filters: AnalyticsFilters) -> CombinedInsightsResponse:
        """Get combined insights across all views"""
        
        product_customer_matrix = self.get_product_customer_matrix(filters, "quote_count")
        top_combinations = self.get_top_product_customer_combinations(filters, 10)
        funnel = self.get_quote_status_funnel(filters)
        velocity = self.get_quote_velocity(12)
        processing_time = self.get_avg_processing_time(filters)
        product_mix = self.get_product_mix_trend(filters)
        
        return CombinedInsightsResponse(
            product_customer_matrix=product_customer_matrix,
            top_combinations=top_combinations,
            funnel=funnel,
            velocity=velocity,
            avg_processing_time=processing_time,
            product_mix_trend=product_mix,
            filters_applied=filters.model_dump(),
            data_timestamp=datetime.now()
        )
    
    def get_product_customer_matrix(self, filters: AnalyticsFilters, metric: str) -> Dict[str, Any]:
        """Get product × customer matrix"""
        
        if metric == "revenue":
            # Get revenue data
            query = self.db.query(
                Project.customer_name,
                TechnicalQuotation.part_type,
                func.coalesce(func.sum(CommercialQuotation.total_amount), 0).label('value')
            ).join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            ).outerjoin(
                CommercialQuotation, Project.quotation_number == CommercialQuotation.quotation_number
            )
        else:
            # Get quote count
            query = self.db.query(
                Project.customer_name,
                TechnicalQuotation.part_type,
                func.count(TechnicalQuotation.id).label('value')
            ).join(
                TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
            )
        
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        
        query = query.group_by(Project.customer_name, TechnicalQuotation.part_type)
        
        results = query.all()
        
        # Organize data
        matrix = {}
        customers = set()
        products = set()
        
        for r in results:
            customer = r.customer_name
            product = self.get_part_type_name(r.part_type)
            value = float(r.value)
            
            customers.add(customer)
            products.add(product)
            
            if customer not in matrix:
                matrix[customer] = {}
            
            matrix[customer][product] = value
        
        # Fill missing values with 0
        for customer in customers:
            for product in products:
                if product not in matrix[customer]:
                    matrix[customer][product] = 0
        
        return {
            "customers": sorted(list(customers)),
            "products": sorted(list(products)),
            "data": matrix,
            "metric": metric
        }
    
    def get_top_product_customer_combinations(self, filters: AnalyticsFilters, limit: int) -> List[Dict[str, Any]]:
        """Get top product-customer combinations"""
        
        query = self.db.query(
            Project.customer_name,
            TechnicalQuotation.part_type,
            func.count(TechnicalQuotation.id).label('quote_count'),
            func.coalesce(func.sum(CommercialQuotation.total_amount), 0).label('total_revenue')
        ).join(
            TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
        ).outerjoin(
            CommercialQuotation, Project.quotation_number == CommercialQuotation.quotation_number
        )
        
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        
        query = query.group_by(Project.customer_name, TechnicalQuotation.part_type)
        query = query.order_by(func.count(TechnicalQuotation.id).desc())
        query = query.limit(limit)
        
        results = query.all()
        
        return [
            {
                "customer": r.customer_name,
                "product": self.get_part_type_name(r.part_type),
                "quote_count": r.quote_count,
                "revenue": round(float(r.total_revenue or 0), 2)
            }
            for r in results
        ]
    
    def get_quote_status_funnel(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get quote status funnel data"""
        
        query = self.db.query(
            Project.quote_status,
            func.count(Project.id).label('count'),
            func.coalesce(func.sum(CommercialQuotation.total_amount), 0).label('value')
        ).outerjoin(
            CommercialQuotation, Project.quotation_number == CommercialQuotation.quotation_number
        )
        
        query = self.apply_date_filter(query, Project, filters)
        
        query = query.group_by(Project.quote_status)
        
        results = query.all()
        
        # Define funnel order
        funnel_order = ['Budgetary', 'Active', 'Won', 'Lost']
        
        funnel_data = {r.quote_status: {'count': r.count, 'value': float(r.value or 0)} for r in results}
        
        funnel = []
        for status in funnel_order:
            if status in funnel_data:
                funnel.append({
                    "stage": status,
                    "count": funnel_data[status]['count'],
                    "value": round(funnel_data[status]['value'], 2)
                })
            else:
                funnel.append({
                    "stage": status,
                    "count": 0,
                    "value": 0
                })
        
        return funnel
    
    def get_quote_velocity(self, months: int) -> List[Dict[str, Any]]:
        """Get monthly quote velocity"""
        
        start_date = datetime.now() - timedelta(days=months * 30)
        
        query = self.db.query(
            func.strftime('%Y-%m', Project.created_at).label('month'),
            func.count(Project.id).label('count')
        ).filter(
            Project.created_at >= start_date
        ).group_by('month').order_by('month')
        
        results = query.all()
        
        return [
            {
                "date": r.month,
                "value": r.count,
                "label": r.month
            }
            for r in results
        ]
    
    def get_avg_processing_time(self, filters: AnalyticsFilters) -> Optional[float]:
        """Get average quote processing time (created to updated)"""
        
        query = self.db.query(
            Project.created_at,
            Project.updated_at
        )
        
        query = self.apply_date_filter(query, Project, filters)
        query = self.apply_status_filter(query, filters)
        
        results = query.all()
        
        if not results:
            return None
        
        time_diffs = []
        for r in results:
            if r.created_at and r.updated_at:
                diff = (r.updated_at - r.created_at).total_seconds() / 3600  # hours
                time_diffs.append(diff)
        
        if not time_diffs:
            return None
        
        return round(sum(time_diffs) / len(time_diffs), 2)
    
    def get_product_mix_trend(self, filters: AnalyticsFilters) -> List[Dict[str, Any]]:
        """Get product mix trend over time"""
        
        query = self.db.query(
            func.strftime('%Y-%m', Project.created_at).label('period'),
            TechnicalQuotation.part_type,
            func.count(TechnicalQuotation.id).label('count')
        ).join(
            TechnicalQuotation, Project.quotation_number == TechnicalQuotation.quotation_number
        )
        
        query = self.apply_date_filter(query, Project, filters)
        
        query = query.group_by('period', TechnicalQuotation.part_type).order_by('period')
        
        results = query.all()
        
        # Organize by period
        mix_data = {}
        products = set()
        
        for r in results:
            period = r.period
            product = self.get_part_type_name(r.part_type)
            products.add(product)
            
            if period not in mix_data:
                mix_data[period] = {'period': period}
            
            mix_data[period][product] = r.count
        
        # Fill missing products
        for period_data in mix_data.values():
            for product in products:
                if product not in period_data:
                    period_data[product] = 0
        
        return sorted(mix_data.values(), key=lambda x: x['period'])
    
    # ========================================================================
    # EXPORT FUNCTIONALITY
    # ========================================================================
    
    def export_analytics_data(self, view: str, format: str, filters: AnalyticsFilters, user_info: Dict[str, str]) -> Dict[str, Any]:
        """Export analytics data to Excel file with professional formatting"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from app.config import EXPORT_DIR

        # Support both CSV and Excel
        if format not in ["csv", "xlsx"]:
            return {"success": False, "error": "Only CSV and Excel exports are supported"}

        try:
            # Get the analytics data
            if view == "product":
                data = self.get_product_analytics(filters)
            elif view == "finance":
                data = self.get_finance_analytics(filters)
            elif view == "customer":
                data = self.get_customer_analytics(filters)
            elif view == "combined":
                data = self.get_combined_insights(filters)
            else:
                return {"success": False, "error": "Invalid view type"}

            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            if format == "xlsx":
                filename = f"{view}_analytics_{timestamp}.xlsx"
                filepath = EXPORT_DIR / filename

                # Create Excel workbook with professional formatting
                wb = Workbook()
                ws = wb.active
                ws.title = f"{view.title()} Analytics"

                # Write Excel content with formatting
                self._write_excel_content(ws, data, view, filters, user_info)

                # Save workbook
                wb.save(filepath)
            else:
                # CSV export (legacy)
                import csv
                filename = f"{view}_analytics_{timestamp}.csv"
                filepath = EXPORT_DIR / filename

                with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    self._write_csv_content(writer, data, view, filters)

            return {
                "success": True,
                "message": f"Data exported successfully to {filename}",
                "filename": filename,
                "filepath": str(filepath),
                "generated_at": datetime.now().isoformat()
            }

        except Exception as e:
            logger.error(f"Export error: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return {"success": False, "error": f"Export failed: {str(e)}"}

    def _write_excel_content(self, ws, data: Any, view: str, filters: AnalyticsFilters, user_info: Dict[str, str]):
        """Write professionally formatted Excel content with colors and styling"""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # Color scheme - Professional Blue & Orange
        HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Blue
        SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Medium Blue
        SECTION_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # Orange
        KPI_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light Green
        DATA_HEADER_FILL = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")  # Light Blue
        ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light Gray

        # Fonts
        TITLE_FONT = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
        HEADER_FONT = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        SECTION_FONT = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        DATA_HEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        NORMAL_FONT = Font(name='Calibri', size=10)
        BOLD_FONT = Font(name='Calibri', size=10, bold=True)

        # Borders
        THIN_BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Center alignment
        CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
        RIGHT = Alignment(horizontal='right', vertical='center')

        # Helper functions
        def get_data_field(data_obj, field_name):
            if isinstance(data_obj, dict):
                return data_obj.get(field_name)
            else:
                return getattr(data_obj, field_name, None)

        def to_dict(item):
            if hasattr(item, 'model_dump'):
                return item.model_dump()
            elif hasattr(item, 'dict'):
                return item.dict()
            elif isinstance(item, dict):
                return item
            return {}

        row = 1

        # ========== COMPANY HEADER ==========
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "RINGSPANN POWER TRANSMISSION INDIA"
        cell.font = TITLE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        row += 1

        # Report Title
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = f"{view.upper()} ANALYTICS EXPORT REPORT"
        cell.font = HEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        row += 1

        # User and Date Information
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = f"Generated By: {user_info.get('name', 'System User')}"
        cell.font = BOLD_FONT
        cell.alignment = LEFT

        ws.merge_cells(f'D{row}:F{row}')
        cell = ws[f'D{row}']
        cell.value = f"Region: {user_info.get('region', 'India')}"
        cell.font = BOLD_FONT
        cell.alignment = RIGHT
        row += 1

        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = f"Generated On: {datetime.now().strftime('%d-%b-%Y %I:%M:%S %p')}"
        cell.font = NORMAL_FONT
        cell.alignment = CENTER
        row += 2  # Extra spacing

        # ========== FILTERS SECTION ==========
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "APPLIED FILTERS"
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = CENTER
        row += 1

        filters_applied = get_data_field(data, 'filters_applied') or {}
        filter_data = [
            ['Date Range', filters_applied.get('date_filter', 'All Time')],
        ]
        if filters_applied.get('start_date'):
            filter_data.append(['Start Date', filters_applied.get('start_date')])
        if filters_applied.get('end_date'):
            filter_data.append(['End Date', filters_applied.get('end_date')])
        if filters_applied.get('quote_status') and filters_applied.get('quote_status') != 'all':
            filter_data.append(['Quote Status', filters_applied.get('quote_status')])
        if filters_applied.get('product_type') and filters_applied.get('product_type') != 'all':
            filter_data.append(['Product Type', filters_applied.get('product_type')])
        if filters_applied.get('customer') and filters_applied.get('customer') != 'all':
            filter_data.append(['Customer', filters_applied.get('customer')])

        for filter_row in filter_data:
            ws[f'A{row}'] = filter_row[0]
            ws[f'A{row}'].font = BOLD_FONT
            ws[f'B{row}'] = filter_row[1]
            ws[f'B{row}'].font = NORMAL_FONT
            row += 1
        row += 1  # Extra spacing

        # ========== KPIs SECTION ==========
        kpis = get_data_field(data, 'kpis') or {}
        if kpis:
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = "KEY PERFORMANCE INDICATORS"
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            cell.alignment = CENTER
            row += 1

            # KPI Headers
            ws[f'A{row}'] = "KPI"
            ws[f'B{row}'] = "Value"
            for col in ['A', 'B']:
                ws[f'{col}{row}'].font = DATA_HEADER_FONT
                ws[f'{col}{row}'].fill = DATA_HEADER_FILL
                ws[f'{col}{row}'].alignment = CENTER
                ws[f'{col}{row}'].border = THIN_BORDER
            row += 1

            kpi_row_start = row
            for key, kpi_data in kpis.items():
                kpi_dict = to_dict(kpi_data)
                value = kpi_dict.get('value', 0)
                fmt = kpi_dict.get('format_type', 'text')

                if fmt == 'currency':
                    formatted_value = f"₹{value:,.2f}"
                elif fmt == 'number':
                    formatted_value = f"{value:,.0f}"
                else:
                    formatted_value = str(value)

                ws[f'A{row}'] = kpi_dict.get('label', key)
                ws[f'B{row}'] = formatted_value

                # Alternating row colors
                if (row - kpi_row_start) % 2 == 0:
                    ws[f'A{row}'].fill = ALT_ROW_FILL
                    ws[f'B{row}'].fill = ALT_ROW_FILL
                else:
                    ws[f'A{row}'].fill = KPI_FILL
                    ws[f'B{row}'].fill = KPI_FILL

                ws[f'A{row}'].font = BOLD_FONT
                ws[f'B{row}'].font = NORMAL_FONT
                ws[f'A{row}'].border = THIN_BORDER
                ws[f'B{row}'].border = THIN_BORDER
                ws[f'A{row}'].alignment = LEFT
                ws[f'B{row}'].alignment = RIGHT
                row += 1
            row += 1  # Extra spacing

        # Continue with view-specific data sections...
        # This will be added in the next part
        row = self._add_view_specific_data(ws, data, view, row, to_dict, get_data_field,
                                           SECTION_FONT, SECTION_FILL, DATA_HEADER_FONT,
                                           DATA_HEADER_FILL, NORMAL_FONT, ALT_ROW_FILL,
                                           THIN_BORDER, CENTER, LEFT, RIGHT)

        # ========== FOOTER ==========
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = f"Total Records: {get_data_field(data, 'total_records') or 'N/A'}"
        cell.font = BOLD_FONT
        cell.alignment = CENTER
        row += 1

        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "--- END OF REPORT ---"
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_view_specific_data(self, ws, data, view, row, to_dict, get_data_field,
                                SECTION_FONT, SECTION_FILL, DATA_HEADER_FONT, DATA_HEADER_FILL,
                                NORMAL_FONT, ALT_ROW_FILL, THIN_BORDER, CENTER, LEFT, RIGHT):
        """Add view-specific data sections to Excel"""
        from openpyxl.styles import Alignment

        # This method will handle the rest of the data export
        # For now, return row to complete the structure
        return row

    def _write_csv_content(self, writer, data: Any, view: str, filters: AnalyticsFilters):
        """Write comprehensive CSV content with multiple sections"""

        try:
            # Helper function to get data from dict or object
            def get_data_field(data_obj, field_name):
                if isinstance(data_obj, dict):
                    return data_obj.get(field_name)
                else:
                    return getattr(data_obj, field_name, None)

            # Helper to convert model to dict
            def to_dict(item):
                if hasattr(item, 'model_dump'):
                    return item.model_dump()
                elif hasattr(item, 'dict'):
                    return item.dict()
                elif isinstance(item, dict):
                    return item
                return {}

            # Add export metadata header
            writer.writerow(['=== RINGSPANN POWER TRANSMISSION INDIA ==='])
            writer.writerow([f'{view.upper()} ANALYTICS EXPORT REPORT'])
            writer.writerow([f'Generated: {datetime.now().strftime("%d-%b-%Y %I:%M %p")}'])
            writer.writerow([])

            # Add filters applied section
            filters_applied = get_data_field(data, 'filters_applied') or {}
            writer.writerow(['=== FILTERS APPLIED ==='])
            writer.writerow(['Date Range', filters_applied.get('date_filter', 'All Time')])
            if filters_applied.get('start_date'):
                writer.writerow(['Start Date', filters_applied.get('start_date')])
            if filters_applied.get('end_date'):
                writer.writerow(['End Date', filters_applied.get('end_date')])
            if filters_applied.get('quote_status') and filters_applied.get('quote_status') != 'all':
                writer.writerow(['Quote Status', filters_applied.get('quote_status')])
            if filters_applied.get('product_type') and filters_applied.get('product_type') != 'all':
                writer.writerow(['Product Type', filters_applied.get('product_type')])
            if filters_applied.get('customer') and filters_applied.get('customer') != 'all':
                writer.writerow(['Customer', filters_applied.get('customer')])
            writer.writerow([])

            if view == "product":
                # KPIs Section
                kpis = get_data_field(data, 'kpis') or {}
                writer.writerow(['=== KEY PERFORMANCE INDICATORS ==='])
                for key, kpi_data in kpis.items():
                    kpi_dict = to_dict(kpi_data)
                    value = kpi_dict.get('value', 0)
                    fmt = kpi_dict.get('format_type', 'text')
                    if fmt == 'currency':
                        formatted_value = f"₹{value:,.2f}"
                    elif fmt == 'number':
                        formatted_value = f"{value:,.0f}"
                    else:
                        formatted_value = str(value)
                    writer.writerow([kpi_dict.get('label', key), formatted_value])
                writer.writerow([])

                # Product Quote Distribution
                quote_dist = get_data_field(data, 'quote_distribution') or []
                if quote_dist:
                    writer.writerow(['=== PRODUCT QUOTE DISTRIBUTION ==='])
                    writer.writerow(['Product Type', 'Quote Count', 'Total Revenue (₹)', 'Avg Value (₹)', 'Percentage'])
                    for item in quote_dist:
                        item_dict = to_dict(item)
                        writer.writerow([
                            item_dict.get('product_type', 'N/A'),
                            item_dict.get('quote_count', 0),
                            f"₹{item_dict.get('revenue', 0):,.2f}",
                            f"₹{item_dict.get('avg_value', 0):,.2f}",
                            f"{item_dict.get('percentage', 0):.1f}%" if item_dict.get('percentage') else 'N/A'
                        ])
                    writer.writerow([])

                # Revenue Contribution
                revenue_contrib = get_data_field(data, 'revenue_contribution') or []
                if revenue_contrib:
                    writer.writerow(['=== REVENUE CONTRIBUTION (WON QUOTES) ==='])
                    writer.writerow(['Product Type', 'Revenue (₹)', 'Percentage'])
                    for item in revenue_contrib:
                        item_dict = to_dict(item)
                        writer.writerow([
                            item_dict.get('product_type', 'N/A'),
                            f"₹{item_dict.get('revenue', 0):,.2f}",
                            f"{item_dict.get('percentage', 0):.1f}%" if item_dict.get('percentage') else 'N/A'
                        ])
                    writer.writerow([])

                # Status Breakdown
                status_breakdown = get_data_field(data, 'status_breakdown') or []
                if status_breakdown:
                    writer.writerow(['=== PRODUCT QUOTE STATUS BREAKDOWN ==='])
                    writer.writerow(['Product Type', 'Budgetary', 'Active', 'Won', 'Lost', 'Total'])
                    for item in status_breakdown:
                        item_dict = to_dict(item) if not isinstance(item, dict) else item
                        total = item_dict.get('Budgetary', 0) + item_dict.get('Active', 0) + item_dict.get('Won', 0) + item_dict.get('Lost', 0)
                        writer.writerow([
                            item_dict.get('product_type', 'N/A'),
                            item_dict.get('Budgetary', 0),
                            item_dict.get('Active', 0),
                            item_dict.get('Won', 0),
                            item_dict.get('Lost', 0),
                            total
                        ])
                    writer.writerow([])

            elif view == "finance":
                # KPIs Section
                kpis = get_data_field(data, 'kpis') or {}
                writer.writerow(['=== KEY PERFORMANCE INDICATORS ==='])
                for key, kpi_data in kpis.items():
                    kpi_dict = to_dict(kpi_data)
                    value = kpi_dict.get('value', 0)
                    fmt = kpi_dict.get('format_type', 'text')
                    if fmt == 'currency':
                        formatted_value = f"₹{value:,.2f}"
                    elif fmt == 'number':
                        formatted_value = f"{value:,.0f}"
                    else:
                        formatted_value = str(value)
                    writer.writerow([kpi_dict.get('label', key), formatted_value])
                writer.writerow([])

                # Revenue by Status
                revenue_data = get_data_field(data, 'revenue_by_status') or []
                if revenue_data:
                    # Convert to dicts if needed
                    revenue_dicts = [to_dict(item) if not isinstance(item, dict) else item for item in revenue_data]
                    total_revenue = sum(item.get('value', 0) for item in revenue_dicts)
                    writer.writerow(['=== REVENUE BY QUOTE STATUS ==='])
                    writer.writerow(['Quote Status', 'Total Revenue (₹)', 'Quote Count', 'Avg Revenue (₹)', 'Percentage'])
                    for item in revenue_dicts:
                        metadata = item.get('metadata', {}) or {}
                        revenue = item.get('value', 0)
                        percentage = (revenue / total_revenue * 100) if total_revenue > 0 else 0
                        label = item.get('label', 'N/A')
                        # Clean up enum format
                        if '.' in str(label):
                            label = str(label).split('.')[-1].title()
                        writer.writerow([
                            label,
                            f"₹{revenue:,.2f}",
                            metadata.get('quote_count', 0),
                            f"₹{metadata.get('avg_revenue', 0):,.2f}",
                            f"{percentage:.1f}%"
                        ])
                    writer.writerow([])

                # Revenue by Product
                product_revenue = get_data_field(data, 'product_revenue') or []
                if product_revenue:
                    writer.writerow(['=== REVENUE BY PRODUCT TYPE ==='])
                    writer.writerow(['Product Type', 'Total Revenue (₹)', 'Quote Count', 'Avg Value (₹)', 'Percentage'])
                    for item in product_revenue:
                        item_dict = to_dict(item) if not isinstance(item, dict) else item
                        writer.writerow([
                            item_dict.get('product_type', 'N/A'),
                            f"₹{item_dict.get('revenue', 0):,.2f}",
                            item_dict.get('quote_count', 0),
                            f"₹{item_dict.get('avg_value', 0):,.2f}",
                            f"{item_dict.get('percentage', 0):.1f}%"
                        ])
                    writer.writerow([])

                # Monthly Trend
                monthly_trend = get_data_field(data, 'monthly_trend') or []
                if monthly_trend:
                    writer.writerow(['=== MONTHLY REVENUE TREND ==='])
                    writer.writerow(['Month', 'Quote Count', 'Total Revenue (₹)', 'Avg Revenue (₹)'])
                    for item in monthly_trend:
                        item_dict = to_dict(item) if not isinstance(item, dict) else item
                        writer.writerow([
                            item_dict.get('month', 'N/A'),
                            item_dict.get('quote_count', 0),
                            f"₹{item_dict.get('total_revenue', 0):,.2f}",
                            f"₹{item_dict.get('avg_revenue', 0):,.2f}"
                        ])

            elif view == "customer":
                # KPIs Section
                kpis = get_data_field(data, 'kpis') or {}
                writer.writerow(['=== KEY PERFORMANCE INDICATORS ==='])
                for key, kpi_data in kpis.items():
                    kpi_dict = to_dict(kpi_data)
                    value = kpi_dict.get('value', 0)
                    fmt = kpi_dict.get('format_type', 'text')
                    if fmt == 'currency':
                        formatted_value = f"₹{value:,.2f}"
                    elif fmt == 'number':
                        formatted_value = f"{value:,.0f}"
                    else:
                        formatted_value = str(value)
                    writer.writerow([kpi_dict.get('label', key), formatted_value])
                writer.writerow([])

                # Top Customers by Revenue
                top_revenue = get_data_field(data, 'top_customers_by_revenue') or []
                if top_revenue:
                    writer.writerow(['=== TOP CUSTOMERS BY REVENUE ==='])
                    writer.writerow(['Rank', 'Customer Name', 'Total Revenue (₹)', 'Quote Count', 'Avg Deal Size (₹)', 'Last Quote Date'])
                    for idx, item in enumerate(top_revenue, 1):
                        item_dict = to_dict(item)
                        writer.writerow([
                            idx,
                            item_dict.get('customer_name', 'N/A'),
                            f"₹{item_dict.get('revenue', 0):,.2f}",
                            item_dict.get('quote_count', 0),
                            f"₹{item_dict.get('avg_deal_size', 0):,.2f}",
                            item_dict.get('last_quote_date', 'N/A')
                        ])
                    writer.writerow([])

                # Top Customers by Quote Count
                top_count = get_data_field(data, 'top_customers_by_count') or []
                if top_count:
                    writer.writerow(['=== TOP CUSTOMERS BY QUOTE COUNT ==='])
                    writer.writerow(['Rank', 'Customer Name', 'Quote Count', 'Total Revenue (₹)', 'Avg Deal Size (₹)'])
                    for idx, item in enumerate(top_count, 1):
                        item_dict = to_dict(item)
                        writer.writerow([
                            idx,
                            item_dict.get('customer_name', 'N/A'),
                            item_dict.get('quote_count', 0),
                            f"₹{item_dict.get('revenue', 0):,.2f}",
                            f"₹{item_dict.get('avg_deal_size', 0):,.2f}"
                        ])
                    writer.writerow([])

                # Customer Status Breakdown
                status_breakdown = get_data_field(data, 'customer_status_breakdown') or []
                if status_breakdown:
                    writer.writerow(['=== CUSTOMER QUOTE STATUS BREAKDOWN ==='])
                    writer.writerow(['Customer Name', 'Budgetary', 'Active', 'Won', 'Lost', 'Total'])
                    for item in status_breakdown:
                        item_dict = to_dict(item) if not isinstance(item, dict) else item
                        writer.writerow([
                            item_dict.get('customer_name', 'N/A'),
                            item_dict.get('Budgetary', 0),
                            item_dict.get('Active', 0),
                            item_dict.get('Won', 0),
                            item_dict.get('Lost', 0),
                            item_dict.get('total', 0)
                        ])

            elif view == "combined":
                # Top Product-Customer Combinations
                top_combos = get_data_field(data, 'top_combinations') or []
                if top_combos:
                    writer.writerow(['=== TOP PRODUCT-CUSTOMER COMBINATIONS ==='])
                    writer.writerow(['Rank', 'Customer', 'Product', 'Quote Count', 'Revenue (₹)'])
                    for idx, item in enumerate(top_combos, 1):
                        item_dict = to_dict(item) if not isinstance(item, dict) else item
                        writer.writerow([
                            idx,
                            item_dict.get('customer', 'N/A'),
                            item_dict.get('product', 'N/A'),
                            item_dict.get('quote_count', 0),
                            f"₹{item_dict.get('revenue', 0):,.2f}"
                        ])
                    writer.writerow([])

                # Quote Status Funnel
                funnel = get_data_field(data, 'funnel') or []
                if funnel:
                    writer.writerow(['=== QUOTE STATUS FUNNEL ==='])
                    writer.writerow(['Stage', 'Quote Count', 'Value (₹)'])
                    for item in funnel:
                        item_dict = to_dict(item) if not isinstance(item, dict) else item
                        writer.writerow([
                            item_dict.get('stage', 'N/A'),
                            item_dict.get('count', 0),
                            f"₹{item_dict.get('value', 0):,.2f}"
                        ])
                    writer.writerow([])

                # Processing Time
                avg_time = get_data_field(data, 'avg_processing_time')
                if avg_time:
                    writer.writerow(['=== QUOTE PROCESSING METRICS ==='])
                    writer.writerow(['Average Processing Time', f'{avg_time:.1f} hours'])
                    writer.writerow([])

            # Add footer
            writer.writerow([])
            writer.writerow(['=== END OF REPORT ==='])
            writer.writerow([f'Total Records: {get_data_field(data, "total_records") or "N/A"}'])
            writer.writerow([f'Report Generated: {datetime.now().strftime("%d-%b-%Y %I:%M:%S %p")}'])

        except Exception as e:
            logger.error(f"Error writing CSV: {e}")
            import traceback
            logger.error(traceback.format_exc())
            raise





































# """
# Analytics Service
# """
# import pandas as pd
# from app.database.connection import SessionLocal, engine
# from datetime import datetime, timedelta

# class AnalyticsService:
#     def get_overview(self):
#         """Get overview analytics"""
#         db = SessionLocal()
#         try:
#             # Use pandas for easy aggregation
#             projects_df = pd.read_sql("SELECT * FROM projects", engine)
            
#             if len(projects_df) == 0:
#                 return {
#                     'total_projects': 0,
#                     'total_customers': 0,
#                     'completion_rate': 0,
#                     'active_projects': 0
#                 }
            
#             return {
#                 'total_projects': len(projects_df),
#                 'total_customers': projects_df['customer_name'].nunique(),
#                 'completion_rate': (projects_df['status'] == 'completed').sum() / len(projects_df) * 100,
#                 'active_projects': (projects_df['status'] == 'in_progress').sum()
#             }
#         finally:
#             db.close()
    
#     def get_product_analytics(self, date_filter, customer_filter):
#         """Get product analytics"""
#         # TODO: Implement with pandas
#         return {
#             'total_types': 4,
#             'products': []
#         }
